"""
@author:Administrator
@file:load_excel.py
@time:2020/05/09
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlrd


class OpenExcel(object):
    def __init__(self, book_name):
        self.book_name = book_name
        self.workbook = self.open_workbook()
        self.wb = load_workbook(filename=self.book_name)
        self.bg_color = PatternFill('solid', fgColor='92d050')  # 设置背景色

    def open_workbook(self):
        return xlrd.open_workbook(self.book_name)

    def get_book(self):
        print('load_excel: get_book: 开始获取表格内容')
        table = self.workbook.sheets()[0]
        item = []
        for rown in range(table.nrows):
            array = {'编号': '', '名称': '', 'url': '', 'type': '', 'rown': ''}
            array['编号'] = table.cell_value(rown, 0)
            array['名称'] = table.cell_value(rown, 1)
            array['url'] = table.cell_value(rown, 3)
            array['type'] = table.cell_value(rown, 6)
            array['rown'] = rown
            item.append(array)
        print('load_excel: get_book: 获取表格内容成功')
        return item

    def get_url(self, data):
        item = []
        print('load_excel: get_url: 开始获取表格url')
        for i in data:
            if i['url'] == '拼多多链接':
                continue
            if i['url'] == '':
                continue
            item.append(i)
        print('load_excel: get_url: 成功获取表格url')
        return item

    def clear_url(self, item):
        print('load_excel: clear_url: 开始清洗表格url')
        item_url = []
        for i in item:
            if i['type'] != '已发布':
                item_url.append(i)
        print('load_excel: clear_url: 清洗表格url成功')
        return item_url

    def modify_excel(self, rown):
        print('load_excel: modify_excel: 开始更改excel信息')
        ws = self.wb['Sheet1']
        row = 'G{}'.format(rown+1)  # 构建url
        ws[row] = '已发布'
        ws[row].fill = self.bg_color
        self.wb.save(self.book_name)
        print('load_excel: modify_excel: 更改excel成功')

    def run(self):
        # 1.打开文件
        # 2.获取表单
        print('load_excel: 打开excel')
        data = self.get_book()
        # 3. 获取url
        print('load_excel: 获取excelurl')
        item = self.get_url(data)
        # 4. 处理只留下未发布的url
        print('load_excel: 清理数据，清理掉已发布的数据')
        item_url = self.clear_url(item)
        print('load_excel: 完成。 清洗表格url 成功')
        return item_url


